"""
test generuje plik .xlsx z wykresem XY funkcji sin(x) 

pre:
    pip install openpyxl
    
"""
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
import numpy as np

wb = Workbook()
ws = wb.active
ws.title = "Sin(x) Data"

ws['A1'] = 'x'
ws['B1'] = 'sin(x)'
len = 50

step = 2*np.pi/(len-1)
ws['A2'] = 0
ws['B2'] = '=SIN(A2)'

for i in range(3, len+2):
    ws[f'A{i}'] = f'=A{i-1}+{step}'
    ws[f'B{i}'] = f'=SIN(A{i})'

chart = ScatterChart()
chart.title = "Wykres funkcji sin(x)"
chart.x_axis.title = 'x'
chart.y_axis.title = 'sin(x)'

xValues = Reference(ws, min_col=1, min_row=2, max_row=len+1)
yValues = Reference(ws, min_col=2, min_row=2, max_row=len+1)
series = Series(yValues, xValues, title="sin(x)")
chart.series.append(series)
ws.add_chart(chart, "D2")

fname = 'excel-sin-chart.xlsx'
wb.save(fname)
print(f"Plik '{fname}' z wykresem został utworzony")